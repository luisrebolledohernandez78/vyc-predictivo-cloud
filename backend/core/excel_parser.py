"""
Parser para archivos Excel de Equipos y Activos
Estructura esperada:
- Columna A: Área (Aserradero, Elaborado, Caldera)
- Columna B: Nombre Equipo
- Columna C: Nombre Activo
- Columna D: Observaciones
"""
import openpyxl
from .models import Area, Equipo, Activo


class ExcelEquiposParser:
    def __init__(self, archivo, sucursal):
        self.archivo = archivo
        self.sucursal = sucursal
        self.errores = []
        self.advertencias = []
        self.datos_parseados = []
    
    def parsear(self):
        """Parsea el archivo Excel y retorna estructura jerárquica"""
        try:
            wb = openpyxl.load_workbook(self.archivo)
            ws = wb.active
            
            # Detectar fila de inicio (la primera con "Aserradero", "Elaborado" o "Caldera")
            inicio_datos = 2  # Default: fila 2
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                if row[0] and str(row[0]).strip().lower() in ['aserradero', 'elaborado', 'caldera']:
                    inicio_datos = idx
                    break
            
            datos = []
            area_actual = None
            equipo_actual = None
            
            # Parsear desde la fila detectada
            for row_idx, row in enumerate(ws.iter_rows(min_row=inicio_datos, values_only=True), start=inicio_datos):
                # Saltar filas completamente vacías
                if all(cell is None or str(cell).strip() == '' for cell in row[:3]):
                    continue
                
                # Procesar celdas y limpiar espacios
                area_nombre = str(row[0]).strip() if row[0] else None
                equipo_nombre = str(row[1]).strip() if row[1] else None
                activo_nombre = str(row[2]).strip() if row[2] else None
                observaciones = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                # Remover valores que son solo espacios o "None"
                if area_nombre == '' or area_nombre == 'None':
                    area_nombre = None
                if equipo_nombre == '' or equipo_nombre == 'None':
                    equipo_nombre = None
                if activo_nombre == '' or activo_nombre == 'None':
                    activo_nombre = None
                
                # Nueva área detectada
                if area_nombre:
                    area_actual = self._normalizar_area(area_nombre)
                    if not area_actual:
                        self.errores.append(f"Fila {row_idx}: Área no válida: {area_nombre}")
                        continue
                    equipo_actual = None  # Reset equipo cuando cambia área
                
                # Nuevo equipo detectado
                if equipo_nombre and area_actual:
                    equipo_actual = equipo_nombre
                
                # Nuevo activo detectado
                if activo_nombre and equipo_actual and area_actual:
                    datos.append({
                        'area': area_actual,
                        'equipo': equipo_actual,
                        'activo': activo_nombre,
                        'observaciones': observaciones if observaciones else None
                    })
            
            self.datos_parseados = datos
            return datos
        
        except Exception as e:
            self.errores.append(f"Error al parsear archivo: {str(e)}")
            return []
    
    def _normalizar_area(self, area_nombre):
        """Normaliza el nombre del área a su valor en choices"""
        if not area_nombre:
            return None
        
        area_lower = str(area_nombre).strip().lower()
        
        mapeo = {
            'aserradero': 'aserradero',
            'elaborado': 'elaborado',
            'caldera': 'caldera',
        }
        
        return mapeo.get(area_lower)
    
    def obtener_preview(self):
        """Retorna un preview de los datos que serán importados con detalle de activos"""
        if not self.datos_parseados:
            return {
                'total_filas': 0,
                'por_area': {},
                'errores': self.errores
            }
        
        # Orden estándar de áreas
        AREA_ORDER = {
            'aserradero': 1,
            'elaborado': 2,
            'caldera': 3,
        }
        
        preview = {
            'total_filas': len(self.datos_parseados),
            'por_area': {},
            'errores': self.errores,
            'advertencias': self.advertencias
        }
        
        for dato in self.datos_parseados:
            area = dato['area']
            if area not in preview['por_area']:
                preview['por_area'][area] = {
                    'equipos': {},
                    'total_activos': 0
                }
            
            equipo = dato['equipo']
            if equipo not in preview['por_area'][area]['equipos']:
                preview['por_area'][area]['equipos'][equipo] = {
                    'activos': [],
                    'cantidad': 0
                }
            
            # Agregar detalle del activo
            preview['por_area'][area]['equipos'][equipo]['activos'].append({
                'nombre': dato['activo'],
                'observaciones': dato['observaciones']
            })
            preview['por_area'][area]['equipos'][equipo]['cantidad'] += 1
            preview['por_area'][area]['total_activos'] += 1
        
        # Ordenar áreas según orden estándar
        preview['por_area'] = dict(sorted(
            preview['por_area'].items(),
            key=lambda x: AREA_ORDER.get(x[0], 999)
        ))
        
        return preview
    
    def importar(self, accion='merge'):
        """
        Importa los datos a la BD
        
        Acciones:
        - 'reemplazar': Elimina todos los equipos/activos existentes en la sucursal
        - 'merge': Agrega nuevos, mantiene existentes (evita duplicados)
        - 'upsert': Actualiza existentes, agrega nuevos
        """
        if not self.datos_parseados:
            self.errores.append("No hay datos para importar")
            return {'exito': False, 'errores': self.errores}
        
        try:
            resultados = {
                'equipos_creados': 0,
                'equipos_actualizados': 0,
                'activos_creados': 0,
                'activos_actualizados': 0,
                'errores': []
            }
            
            # Si es reemplazar, elimina primero
            if accion == 'reemplazar':
                Equipo.objects.filter(area__sucursal=self.sucursal).delete()
                resultados['nota'] = 'Equipos y activos anteriores eliminados'
            
            # Procesar datos
            for dato in self.datos_parseados:
                try:
                    # Obtener o crear área
                    area = Area.objects.get(
                        sucursal=self.sucursal,
                        nombre=dato['area']
                    )
                    
                    # Obtener o crear equipo
                    if accion == 'upsert':
                        equipo, creado = Equipo.objects.update_or_create(
                            area=area,
                            nombre=dato['equipo'],
                            defaults={'observaciones': dato['observaciones'] or ''}
                        )
                        if creado:
                            resultados['equipos_creados'] += 1
                        else:
                            resultados['equipos_actualizados'] += 1
                    else:  # merge
                        equipo, creado = Equipo.objects.get_or_create(
                            area=area,
                            nombre=dato['equipo'],
                            defaults={'observaciones': dato['observaciones'] or ''}
                        )
                        if creado:
                            resultados['equipos_creados'] += 1
                    
                    # Obtener o crear activo
                    if accion == 'upsert':
                        activo, creado = Activo.objects.update_or_create(
                            equipo=equipo,
                            nombre=dato['activo'],
                            defaults={'observaciones': dato['observaciones'] or ''}
                        )
                        if creado:
                            resultados['activos_creados'] += 1
                        else:
                            resultados['activos_actualizados'] += 1
                    else:  # merge
                        activo, creado = Activo.objects.get_or_create(
                            equipo=equipo,
                            nombre=dato['activo'],
                            defaults={'observaciones': dato['observaciones'] or ''}
                        )
                        if creado:
                            resultados['activos_creados'] += 1
                
                except Area.DoesNotExist:
                    resultados['errores'].append(
                        f"Área '{dato['area']}' no existe en la sucursal"
                    )
                except Exception as e:
                    resultados['errores'].append(
                        f"Error al importar equipo '{dato['equipo']}': {str(e)}"
                    )
            
            resultados['exito'] = True
            return resultados
        
        except Exception as e:
            return {
                'exito': False,
                'errores': [f"Error general en importación: {str(e)}"]
            }
